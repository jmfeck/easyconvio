from __future__ import annotations

import csv
import os
import shutil
import subprocess
from typing import Optional, List, Any, Iterator

from .base import BaseFile


def _require_command(cmd: str, install_hint: str) -> None:
    if shutil.which(cmd) is None:
        raise RuntimeError(f"'{cmd}' not found. {install_hint}")


def _libreoffice_binary() -> str:
    """Return the LibreOffice CLI name (`libreoffice` on Linux, `soffice` on Windows)."""
    for name in ("libreoffice", "soffice"):
        if shutil.which(name):
            return name
    raise RuntimeError(
        "LibreOffice not found. Install LibreOffice: https://www.libreoffice.org/"
    )


_LIBREOFFICE_FORMATS = {"xls", "xlsx", "ods"}
_DELIMITED_FORMATS = {"csv", "tsv"}


class SpreadsheetFile(BaseFile):
    """Spreadsheet with read/write across xlsx, xls, ods, csv, tsv.

    XLSX/ODS use openpyxl/odfpy directly. XLS reads via xlrd or LibreOffice
    fallback. CSV/TSV use the stdlib csv module. Conversions between groups
    go through an intermediate row representation.
    """

    def _load(self) -> None:
        self._sheets: dict[str, list[list[Any]]] = {}
        fmt = self.format
        if fmt in _DELIMITED_FORMATS:
            self._sheets["Sheet1"] = self._read_delimited(self.path, fmt)
        elif fmt == "xlsx":
            self._sheets = self._read_xlsx(self.path)
        elif fmt == "ods":
            self._sheets = self._read_ods(self.path)
        elif fmt == "xls":
            self._sheets = self._read_xls(self.path)
        else:
            raise ValueError(f"Unsupported spreadsheet format: {fmt}")

    # --- Properties ---

    @property
    def sheet_names(self) -> List[str]:
        """Names of sheets in order."""
        return list(self._sheets.keys())

    @property
    def sheet_count(self) -> int:
        """Number of sheets."""
        return len(self._sheets)

    def rows(self, sheet: Optional[str] = None) -> List[List[Any]]:
        """Rows from the named sheet (or first sheet by default)."""
        name = sheet or self.sheet_names[0]
        return self._sheets[name]

    def row_count(self, sheet: Optional[str] = None) -> int:
        """Number of rows in the named sheet."""
        return len(self.rows(sheet))

    # --- Readers ---

    @staticmethod
    def _read_delimited(path: str, fmt: str) -> List[List[str]]:
        delim = "\t" if fmt == "tsv" else ","
        with open(path, newline="", encoding="utf-8") as f:
            return [list(row) for row in csv.reader(f, delimiter=delim)]

    @staticmethod
    def _read_xlsx(path: str) -> dict[str, list[list[Any]]]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "XLSX support requires openpyxl. "
                "Install with: pip install easyconvio[spreadsheets]"
            )
        wb = load_workbook(path, data_only=True)
        return {
            name: [list(r) for r in wb[name].iter_rows(values_only=True)]
            for name in wb.sheetnames
        }

    @staticmethod
    def _read_ods(path: str) -> dict[str, list[list[Any]]]:
        try:
            from odf.opendocument import load
            from odf.table import Table, TableRow, TableCell
            from odf.text import P
        except ImportError:
            raise ImportError(
                "ODS support requires odfpy. "
                "Install with: pip install easyconvio[spreadsheets]"
            )
        doc = load(path)
        sheets: dict[str, list[list[Any]]] = {}
        for table in doc.spreadsheet.getElementsByType(Table):
            name = table.getAttribute("name")
            rows: list[list[Any]] = []
            for row_el in table.getElementsByType(TableRow):
                row: list[Any] = []
                for cell in row_el.getElementsByType(TableCell):
                    repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
                    text = "".join(str(p) for p in cell.getElementsByType(P))
                    row.extend([text] * repeat)
                rows.append(row)
            sheets[name] = rows
        return sheets

    @staticmethod
    def _read_xls(path: str) -> dict[str, list[list[Any]]]:
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "XLS support requires xlrd. "
                "Install with: pip install easyconvio[spreadsheets]"
            )
        book = xlrd.open_workbook(path)
        return {
            sheet.name: [sheet.row_values(r) for r in range(sheet.nrows)]
            for sheet in book.sheets()
        }

    # --- Writers ---

    def _write_delimited(self, output_path: str, fmt: str) -> str:
        delim = "\t" if fmt == "tsv" else ","
        rows = self.rows()
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=delim)
            for row in rows:
                writer.writerow(["" if v is None else v for v in row])
        return output_path

    def _write_xlsx(self, output_path: str) -> str:
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ImportError(
                "XLSX support requires openpyxl. "
                "Install with: pip install easyconvio[spreadsheets]"
            )
        wb = Workbook()
        wb.remove(wb.active)
        for name, rows in self._sheets.items():
            ws = wb.create_sheet(title=name[:31])
            for row in rows:
                ws.append(list(row))
        wb.save(output_path)
        return output_path

    def _write_ods(self, output_path: str) -> str:
        try:
            from odf.opendocument import OpenDocumentSpreadsheet
            from odf.table import Table, TableRow, TableCell
            from odf.text import P
        except ImportError:
            raise ImportError(
                "ODS support requires odfpy. "
                "Install with: pip install easyconvio[spreadsheets]"
            )
        doc = OpenDocumentSpreadsheet()
        for name, rows in self._sheets.items():
            table = Table(name=name)
            for row in rows:
                tr = TableRow()
                for value in row:
                    cell = TableCell()
                    cell.addElement(P(text="" if value is None else str(value)))
                    tr.addElement(cell)
                table.addElement(tr)
            doc.spreadsheet.addElement(table)
        doc.save(output_path)
        return output_path

    # --- Operations ---

    def add_sheet(self, name: str, rows: List[List[Any]]) -> SpreadsheetFile:
        """Add a new sheet with the given rows."""
        self._sheets[name] = [list(r) for r in rows]
        return self

    def remove_sheet(self, name: str) -> SpreadsheetFile:
        """Remove a sheet by name."""
        del self._sheets[name]
        return self

    def rename_sheet(self, old: str, new: str) -> SpreadsheetFile:
        """Rename a sheet."""
        self._sheets = {new if k == old else k: v for k, v in self._sheets.items()}
        return self

    def iter_rows(self, sheet: Optional[str] = None) -> Iterator[List[Any]]:
        """Iterate over rows of the named sheet (or first)."""
        for row in self.rows(sheet):
            yield row

    # --- Export ---

    def to_csv(self, output_path: Optional[str] = None) -> str:
        """Export the first sheet as CSV."""
        return self._write_delimited(self._output_path("csv", output_path), "csv")

    def to_tsv(self, output_path: Optional[str] = None) -> str:
        """Export the first sheet as TSV."""
        return self._write_delimited(self._output_path("tsv", output_path), "tsv")

    def to_xlsx(self, output_path: Optional[str] = None) -> str:
        """Export as XLSX (preserves all sheets)."""
        return self._write_xlsx(self._output_path("xlsx", output_path))

    def to_ods(self, output_path: Optional[str] = None) -> str:
        """Export as ODS (preserves all sheets)."""
        return self._write_ods(self._output_path("ods", output_path))

    def to_xls(self, output_path: Optional[str] = None) -> str:
        """Export as legacy XLS via LibreOffice (writing native XLS in Python is unmaintained)."""
        binary = _libreoffice_binary()
        output_path = self._output_path("xls", output_path)
        # Round-trip: write a temp xlsx, convert to xls.
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            xlsx_path = os.path.join(tmp, "tmp.xlsx")
            self._write_xlsx(xlsx_path)
            out_dir = os.path.dirname(os.path.abspath(output_path))
            subprocess.run(
                [binary, "--headless", "--convert-to", "xls",
                 "--outdir", out_dir, xlsx_path],
                check=True,
                capture_output=True,
            )
            generated = os.path.join(out_dir, "tmp.xls")
            if generated != os.path.abspath(output_path):
                os.replace(generated, output_path)
        return output_path
